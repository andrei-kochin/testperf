import os
import sys
import re
import subprocess
import platform
from time import perf_counter
from typing import List, Optional, Tuple


migx_binary = "migraphx-driver.exe" if platform.system() == "Windows" else "migraphx-driver"


def _parse_csv_ints(raw: str) -> List[int]:
  return [int(x.strip()) for x in raw.split(",") if x.strip()]


def _parse_target(argv: List[str]) -> Tuple[str, bool]:
  """
  Returns (target, is_forced).

  - If user passes `--target <gpu|cpu|ref>`, it is considered forced (no fallback).
  - Otherwise defaults to "gpu" and is not forced (we may fallback if GPU isn't available).
  """
  if "--target" not in argv:
    return ("gpu", False)

  i = argv.index("--target")
  if i + 1 >= len(argv):
    return ("gpu", False)

  raw = (argv[i + 1] or "").strip().lower()
  if raw in ("gpu", "cpu", "ref"):
    return (raw, True)

  return ("gpu", False)


def _parse_imgsz(argv: List[str]) -> Tuple[int, int]:
  if "--imgsz" not in argv:
    return (640, 640)

  i = argv.index("--imgsz")
  if i + 1 >= len(argv):
    return (640, 640)

  token = argv[i + 1]
  # Allow: 640x640, 640,640, or "--imgsz 640 640"
  if ("x" in token) or ("," in token):
    sep = "x" if "x" in token else ","
    parts = [p.strip() for p in token.split(sep) if p.strip()]
    if len(parts) == 2:
      return (int(parts[0]), int(parts[1]))
    return (640, 640)

  if i + 2 < len(argv):
    try:
      return (int(argv[i + 1]), int(argv[i + 2]))
    except Exception:
      return (640, 640)

  return (640, 640)


def _round_up_to_stride(x: int, stride: int = 32) -> int:
  if x <= 0:
    return x
  return ((x + stride - 1) // stride) * stride


def _looks_like_no_device_error(text: str) -> bool:
  s = (text or "").lower()
  # Seen in the wild (your log):
  #   get_device_id: No device
  #   ... device_name.cpp ... No device
  return ("no device" in s) or ("get_device_id" in s) or ("device_name.cpp" in s)


def _missing_migraphx_backend_lib(text: str) -> Optional[str]:
  """
  Detect missing MIGraphX backend shared libraries from stderr/stdout text.
  Returns the missing library filename (e.g. "libmigraphx_cpu.so") or None.
  """
  s = (text or "")
  # Typical dlopen error:
  #   ... libmigraphx_cpu.so: cannot open shared object file: No such file or directory
  m = re.search(r"(libmigraphx_(?:cpu|gpu|ref)\.so)[^\\n]*cannot open shared object file", s, re.IGNORECASE)
  if m:
    return m.group(1)
  return None


def get_migraphx_driver_version() -> str:
  try:
    result = subprocess.run([migx_binary, "-v"], capture_output=True, text=True, timeout=10)
    version_output = (result.stdout or "").strip() + (result.stderr or "").strip()
    match = re.search(r"MIGraphX Version:\s*([\d\.\-\w]+)", version_output)
    if match:
      return match.group(1)
    return version_output or "Unknown"
  except Exception as e:
    return f"Unknown (Error: {e})"


def parse_migraphx_perf_output(output: str) -> dict:
  """Parse migraphx-driver perf output and extract performance metrics."""
  data: dict = {}

  batch_match = re.search(r"Batch size:\s*(\d+)", output)
  if batch_match:
    data["batch_size"] = int(batch_match.group(1))

  rate_match = re.search(r"Rate:\s*([\d\.]+)\s*inferences/sec", output)
  if rate_match:
    data["rate"] = float(rate_match.group(1))

  total_time_match = re.search(
    r"Total time:\s*([\d\.]+)ms\s*\(Min:\s*([\d\.]+)ms,\s*Max:\s*([\d\.]+)ms,\s*Mean:\s*([\d\.]+)ms,\s*Median:\s*([\d\.]+)ms\)",
    output,
  )
  if total_time_match:
    data["total_time"] = float(total_time_match.group(1)) / 1000
    data["min_time"] = float(total_time_match.group(2)) / 1000
    data["max_time"] = float(total_time_match.group(3)) / 1000
    data["mean_time"] = float(total_time_match.group(4)) / 1000
    data["median_time"] = float(total_time_match.group(5)) / 1000

  percentiles_match = re.search(
    r"Percentiles\s*\(90%,\s*95%,\s*99%\):\s*\(([\d\.]+)ms,\s*([\d\.]+)ms,\s*([\d\.]+)ms\)",
    output,
  )
  if percentiles_match:
    data["p90"] = float(percentiles_match.group(1)) / 1000
    data["p95"] = float(percentiles_match.group(2)) / 1000
    data["p99"] = float(percentiles_match.group(3)) / 1000

  instructions_match = re.search(r"Total instructions time:\s*([\d\.]+)ms", output)
  if instructions_match:
    data["instructions_time"] = float(instructions_match.group(1)) / 1000

  overhead_match = re.search(r"Overhead time:\s*([\d\.\-]+)ms,\s*([\d\.\-]+)ms", output)
  if overhead_match:
    data["overhead_time1"] = float(overhead_match.group(1)) / 1000
    data["overhead_time2"] = float(overhead_match.group(2)) / 1000

  return data


def ensure_onnx_exported(
  *,
  onnx_path: str,
  batch: int,
  imgsz: Tuple[int, int],
  weights_path: str,
  weights_url: str,
  export_half: bool,
  export_dynamic: bool,
) -> None:
  if os.path.exists(onnx_path):
    return

  if not os.path.exists(weights_path):
    try:
      import urllib.request
      urllib.request.urlretrieve(weights_url, weights_path)
    except Exception as e:
      raise Exception(f"Failed to download weights '{weights_path}' from '{weights_url}': {e}")

  from ultralytics import YOLO

  model = YOLO(weights_path)
  exported_path = model.export(
    format="onnx",
    dynamic=export_dynamic,
    imgsz=[int(imgsz[0]), int(imgsz[1])],
    batch=int(batch),
    half=bool(export_half),
  )

  exported_path = str(exported_path)
  if os.path.abspath(exported_path) != os.path.abspath(onnx_path):
    try:
      os.replace(exported_path, onnx_path)
    except Exception:
      # Best-effort fallback if export produced a different name in same dir
      if os.path.exists(onnx_path):
        return
      raise


def fix_onnx_input_shape_inplace(onnx_path: str, batch: int, imgsz: Tuple[int, int]) -> None:
  """
  Make the first ONNX graph input fixed shape: [batch, 3, H, W].
  Uses pure onnx (no onnxruntime dependency).
  """
  try:
    import onnx
  except Exception:
    return

  model = onnx.load(onnx_path)
  if not model.graph.input:
    return

  inp = model.graph.input[0]
  dims = inp.type.tensor_type.shape.dim
  target = [int(batch), 3, int(imgsz[0]), int(imgsz[1])]
  if len(dims) < 4:
    return

  for i, v in enumerate(target[: len(dims)]):
    dims[i].dim_param = ""
    dims[i].dim_value = int(v)

  onnx.save(model, onnx_path)


def compile_to_mxr(
  *,
  onnx_path: str,
  mxr_path: str,
  quantize_fp16: bool,
  force_driver: bool,
  target: str,
  allow_fallback: bool,
) -> str:
  """
  Compile ONNX -> MXR.
  Prefers Python MIGraphX (for FP16 quantization like the gist) if available, unless forced.
  Returns a string like "python(gpu)" or "driver(cpu)".
  """
  target = (target or "gpu").strip().lower()
  if target not in ("gpu", "cpu", "ref"):
    target = "gpu"

  target_order = [target]
  if allow_fallback:
    for t in ("gpu", "cpu", "ref"):
      if t not in target_order:
        target_order.append(t)

  if not force_driver:
    try:
      import migraphx  # type: ignore

      prog = migraphx.parse_onnx(onnx_path)
      if quantize_fp16:
        try:
          migraphx.quantize_fp16(prog)
        except Exception:
          # Not all builds expose quantize_fp16; continue without it.
          pass

      last_exc: Optional[BaseException] = None
      for idx, t in enumerate(target_order):
        try:
          try:
            prog.compile(t=migraphx.get_target(t), offload_copy=False)
          except TypeError:
            # Older python bindings don't take `t=...`/`offload_copy` kwargs.
            prog.compile(migraphx.get_target(t))

          migraphx.save(prog, mxr_path)
          return f"python({t})"
        except Exception as e:
          last_exc = e
          # Only fall back from GPU -> CPU/REF when it looks like the GPU is not present.
          # (Otherwise we risk masking real GPU compilation bugs by switching targets.)
          if (idx == 0) and (t == "gpu") and allow_fallback and _looks_like_no_device_error(str(e)):
            continue
          # Fall back from CPU -> REF if enabled, otherwise stop.
          if (idx == 0) and (t == "cpu") and allow_fallback:
            continue
          break

      if last_exc is not None:
        raise last_exc
    except Exception:
      pass

  def _driver_compile(cmd_target: str) -> subprocess.CompletedProcess[str]:
    flag = f"--{cmd_target}"
    compile_cmd = [migx_binary, "compile", onnx_path, flag, "--binary", "-o", mxr_path]
    return subprocess.run(compile_cmd, capture_output=True, text=True)

  last_err: str = ""
  for idx, t in enumerate(target_order):
    result = _driver_compile(t)
    if result.returncode == 0:
      return f"driver({t})"
    last_err = (result.stderr or "") + (result.stdout or "")

    # If GPU isn't available, it's common to fail-fast; CPU/ref fallback can still work.
    if (idx == 0) and (t == "gpu"):
      if allow_fallback and _looks_like_no_device_error(last_err):
        continue
      break

    # If CPU compile fails, try ref as a last resort (when fallback is enabled).
    # Note: CPU is often a fallback target after GPU; in that common case idx != 0,
    # so we still want to continue to REF when enabled.
    if (t == "cpu") and allow_fallback and ("ref" in target_order[idx + 1 :]):
      continue

    break

  missing_lib = _missing_migraphx_backend_lib(last_err)
  if missing_lib:
    raise Exception(
      "Failed to compile model via migraphx-driver because a MIGraphX backend library "
      f"could not be loaded: {missing_lib}. "
      "This usually means MIGraphX runtime libraries are not installed (or not on the dynamic "
      "loader search path). On ROCm Linux, install MIGraphX (e.g. `sudo apt install -y migraphx`) "
      "and ensure `/opt/rocm/lib` is discoverable (e.g. via `ldconfig` or `LD_LIBRARY_PATH`). "
      f"Details: {last_err.strip()}"
    )

  if _looks_like_no_device_error(last_err):
    raise Exception(
      "Failed to compile model via migraphx-driver (no GPU device visible). "
      "Try `--target cpu` (or install/configure ROCm/HIP so a GPU is visible). "
      f"Details: {last_err.strip()}"
    )

  raise Exception(f"Failed to compile model via migraphx-driver: {last_err.strip()}")


def run_perf(mxr_path: str) -> str:
  # Try current repo's convention first.
  perf_cmd = [migx_binary, "perf", "--migraphx", mxr_path]
  result = subprocess.run(perf_cmd, capture_output=True, text=True)

  # Some versions use: `migraphx-driver perf model.mxr`
  if (result.returncode != 0) and ("--migraphx" in (result.stderr or "")):
    perf_cmd = [migx_binary, "perf", mxr_path]
    result = subprocess.run(perf_cmd, capture_output=True, text=True)

  if (result.returncode != 0) and ("MIGraphX program was likely compiled with offload_copy set" in (result.stderr or "")):
    perf_cmd.append("--enable-offload-copy")
    result = subprocess.run(perf_cmd, capture_output=True, text=True)

  if result.returncode != 0:
    raise Exception(f"Failed to run perf: {result.stderr}")

  return (result.stdout or "") + (result.stderr or "")


def main(argv: Optional[List[str]] = None, *, default_fp16: bool = False) -> int:
  argv = list(sys.argv[1:] if argv is None else argv)

  if not os.path.exists("./temp"):
    os.makedirs("./temp")

  batches = [1]
  if "--batch-size" in argv:
    try:
      batches = _parse_csv_ints(argv[argv.index("--batch-size") + 1])
    except Exception as e:
      print(f'{{ "Error": "Failed to set batch size {e}, using default [{", ".join(map(str, batches))}]" }},')

  imgsz = _parse_imgsz(argv)
  imgsz = (_round_up_to_stride(imgsz[0], 32), _round_up_to_stride(imgsz[1], 32))

  export_dynamic = "--no-dynamic" not in argv
  export_half = ("--export-half" in argv) or default_fp16 or ("--fp16" in argv)
  quantize_fp16 = ("--quantize-fp16" in argv) or default_fp16 or ("--fp16" in argv)
  force_driver = "--compile-with-driver" in argv
  dry_run = "--dry-run" in argv
  target, target_forced = _parse_target(argv)
  allow_fallback = not target_forced

  weights_path = "yolov11l.pt"
  weights_url = "https://github.com/ultralytics/assets/releases/download/v8.4.0/yolo11l.pt"

  migraphx_version = get_migraphx_driver_version()
  print(f'{{ "MIGraphX Version": "{migraphx_version}" }},')
  print(
    f'{{ "Batches": {batches}, "imgsz": [{imgsz[0]}, {imgsz[1]}], '
    f'"export_half": {export_half}, "quantize_fp16": {quantize_fp16}, '
    f'"dynamic": {export_dynamic}, "force_driver_compile": {force_driver}, '
    f'"target": "{target}", "auto_fallback": {allow_fallback} }},'
  )

  if dry_run:
    print('{ "Status": "DryRun" }')
    return 0

  inference_times = {}
  compile_times = {}
  all_results = {}

  for batch in batches:
    print(f'{{ "Processing Batch": {batch} }},')

    onnx_name = f"yolov11l_{batch}b_{imgsz[0]}x{imgsz[1]}.onnx"
    onnx_path = os.path.join("./temp", onnx_name)
    mxr_path = os.path.join("./temp", onnx_name[:-4] + "mxr")

    try:
      if not os.path.exists(onnx_path):
        print(f'{{ "Exporting Model": "{onnx_path}" }},')
      ensure_onnx_exported(
        onnx_path=onnx_path,
        batch=batch,
        imgsz=imgsz,
        weights_path=weights_path,
        weights_url=weights_url,
        export_half=export_half,
        export_dynamic=export_dynamic,
      )

      # Fix input to a deterministic shape (like the gist) so compile/perf are stable.
      fix_onnx_input_shape_inplace(onnx_path, batch, imgsz)
    except Exception as e:
      print(f'{{ "Error": "Failed to export/fix ONNX {e}" }},')
      continue

    if not os.path.exists(mxr_path):
      print(f'{{ "Compiling MXR": "{mxr_path}" }},')
      try:
        start_time = perf_counter()
        compiler = compile_to_mxr(
          onnx_path=onnx_path,
          mxr_path=mxr_path,
          quantize_fp16=quantize_fp16,
          force_driver=force_driver,
          target=target,
          allow_fallback=allow_fallback,
        )
        compile_time = perf_counter() - start_time
        print(f'{{ "Compile Time": {compile_time}, "Compiler": "{compiler}" }},')
        compile_times[batch] = compile_time
      except Exception as e:
        print(f'{{ "Error": "Failed to compile model {e}" }},')
        continue
    else:
      compile_times[batch] = 0

    print(f'{{ "Running Performance Test": "{mxr_path}" }},')
    try:
      perf_text = run_perf(mxr_path)
      perf_data = parse_migraphx_perf_output(perf_text)
      all_results[batch] = perf_data
      print(f'{{ "Performance Data": {perf_data} }},')

      if "mean_time" in perf_data:
        num_runs = 10
        inference_times[batch] = [perf_data["mean_time"]] * num_runs
        inference_times[batch].append(
          {
            "Minimum": perf_data.get("min_time", perf_data["mean_time"]),
            "Maximum": perf_data.get("max_time", perf_data["mean_time"]),
            "Average": perf_data["mean_time"],
          }
        )
    except Exception as e:
      print(f'{{ "Error": "Failed to run performance test {e}" }},')
      continue

  if inference_times:
    print('{ "Generating Report": "Starting" },')
    try:
      import openpyxl
      from openpyxl.chart import LineChart, Reference, Series
      from openpyxl.chart.layout import Layout, ManualLayout
      from openpyxl.utils import get_column_letter
      from copy import deepcopy
      import datetime

      wb = openpyxl.Workbook()
      main_sheet = wb.active
      main_sheet.title = "Overview"
      inference_sheet = wb.create_sheet("Inference")

      inference_sheet.append(["Inference times (MIGraphX)"])
      inference_sheet.append(["Metric"] + [f"Batch {batch}" for batch in batches])

      offset_col = 2
      offset_stat_row = inference_sheet.max_row + 1

      inference_sheet.append(["Average"])
      inference_sheet.append(["Median"])
      inference_sheet.append(["90th Percentile"])
      inference_sheet.append(["95th Percentile"])
      inference_sheet.append(["99th Percentile"])
      inference_sheet.append(["Minimum"])
      inference_sheet.append(["Maximum"])
      inference_sheet.append(["IPS (Average)"])
      inference_sheet.append(["IPS (Median)"])
      inference_sheet.append(["IPS (90th Percentile)"])
      inference_sheet.append(["IPS (95th Percentile)"])
      inference_sheet.append(["IPS (99th Percentile)"])
      inference_sheet.append(["BPS (Average)"])
      inference_sheet.append(["BPS (Median)"])
      inference_sheet.append(["BPS (90th Percentile)"])
      inference_sheet.append(["BPS (95th Percentile)"])
      inference_sheet.append(["BPS (99th Percentile)"])
      inference_sheet.append(["Compile Time"])

      for batch_index, batch in enumerate(batches):
        col_letter = get_column_letter(offset_col + batch_index)
        if batch in all_results:
          data = all_results[batch]
          inference_sheet[col_letter + str(offset_stat_row + 0)] = data.get("mean_time", 0)
          inference_sheet[col_letter + str(offset_stat_row + 1)] = data.get("median_time", 0)
          inference_sheet[col_letter + str(offset_stat_row + 2)] = data.get("p90", 0)
          inference_sheet[col_letter + str(offset_stat_row + 3)] = data.get("p95", 0)
          inference_sheet[col_letter + str(offset_stat_row + 4)] = data.get("p99", 0)
          inference_sheet[col_letter + str(offset_stat_row + 5)] = data.get("min_time", 0)
          inference_sheet[col_letter + str(offset_stat_row + 6)] = data.get("max_time", 0)

          if data.get("mean_time", 0) > 0:
            inference_sheet[col_letter + str(offset_stat_row + 7)] = 1 / data["mean_time"]
          if data.get("median_time", 0) > 0:
            inference_sheet[col_letter + str(offset_stat_row + 8)] = 1 / data["median_time"]
          if data.get("p90", 0) > 0:
            inference_sheet[col_letter + str(offset_stat_row + 9)] = 1 / data["p90"]
          if data.get("p95", 0) > 0:
            inference_sheet[col_letter + str(offset_stat_row + 10)] = 1 / data["p95"]
          if data.get("p99", 0) > 0:
            inference_sheet[col_letter + str(offset_stat_row + 11)] = 1 / data["p99"]

          inference_sheet[col_letter + str(offset_stat_row + 12)] = f"={batch} * {col_letter + str(offset_stat_row + 7)}"
          inference_sheet[col_letter + str(offset_stat_row + 13)] = f"={batch} * {col_letter + str(offset_stat_row + 8)}"
          inference_sheet[col_letter + str(offset_stat_row + 14)] = f"={batch} * {col_letter + str(offset_stat_row + 9)}"
          inference_sheet[col_letter + str(offset_stat_row + 15)] = f"={batch} * {col_letter + str(offset_stat_row + 10)}"
          inference_sheet[col_letter + str(offset_stat_row + 16)] = f"={batch} * {col_letter + str(offset_stat_row + 11)}"

          inference_sheet[col_letter + str(offset_stat_row + 17)] = compile_times.get(batch, 0)

      chart = LineChart()
      chart.title = "Metrics (Time)"
      chart.x_axis.title = "Batch Size"
      chart.y_axis.title = "Time (s)"
      chart.x_axis.delete = False
      chart.y_axis.delete = False
      metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile", "Minimum", "Maximum"]
      for metric_index in range(len(metrics)):
        series = Series(
          values=Reference(
            inference_sheet,
            min_col=offset_col,
            min_row=offset_stat_row + metric_index,
            max_col=offset_col + len(batches) - 1,
            max_row=offset_stat_row + metric_index,
          ),
          title=f"{metrics[metric_index]}",
        )
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
      batch_titles = Reference(
        inference_sheet,
        min_col=offset_col,
        min_row=offset_stat_row - 1,
        max_col=offset_col + len(batches) - 1,
        max_row=offset_stat_row - 1,
      )
      chart.set_categories(batch_titles)
      chart.legend.position = "b"
      chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.02, h=0.65, w=0.9))
      chart.width = 25
      inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "1")
      main_sheet.add_chart(deepcopy(chart), "F5")

      chart = LineChart()
      chart.title = "IPS (Inferences Per Second)"
      chart.x_axis.title = "Batch Size"
      chart.y_axis.title = "Inferences Per Second"
      chart.x_axis.delete = False
      chart.y_axis.delete = False
      metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile"]
      for metric_index in range(len(metrics)):
        series = Series(
          values=Reference(
            inference_sheet,
            min_col=offset_col,
            min_row=offset_stat_row + 7 + metric_index,
            max_col=offset_col + len(batches) - 1,
            max_row=offset_stat_row + 7 + metric_index,
          ),
          title=f"{metrics[metric_index]}",
        )
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
      chart.set_categories(batch_titles)
      chart.legend.position = "b"
      chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.02, h=0.65, w=0.9))
      chart.width = 15
      inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "16")
      main_sheet.add_chart(deepcopy(chart), "F20")

      chart = LineChart()
      chart.title = "BPS (Batches Per Second)"
      chart.x_axis.title = "Batch Size"
      chart.y_axis.title = "Batches Per Second"
      chart.x_axis.delete = False
      chart.y_axis.delete = False
      for metric_index in range(len(metrics)):
        series = Series(
          values=Reference(
            inference_sheet,
            min_col=offset_col,
            min_row=offset_stat_row + 12 + metric_index,
            max_col=offset_col + len(batches) - 1,
            max_row=offset_stat_row + 12 + metric_index,
          ),
          title=f"{metrics[metric_index]}",
        )
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
      chart.set_categories(batch_titles)
      chart.legend.position = "b"
      chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.02, h=0.65, w=0.9))
      chart.width = 15
      inference_sheet.add_chart(chart, get_column_letter(len(batches) + 11) + "16")
      main_sheet.add_chart(deepcopy(chart), "P20")

      report_datetime = datetime.datetime.now()
      main_sheet.column_dimensions[get_column_letter(1)].width = 30
      main_sheet.append(["Model:", "yolov11l (MIGraphX gist-style)"])
      main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
      main_sheet.append(["Description:", "YOLO11L export/compile/perf pipeline inspired by the MIGraphX gist"])
      main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
      main_sheet.append(["Run Command:", " ".join([sys.executable, "-m", __package__ + ".migx_driver_cache_gist"] + argv)])
      main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
      main_sheet.append(["Report Date:", report_datetime.strftime("%Y-%m-%d %H:%M:%S")])
      main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=6)
      main_sheet.append(["Batches:", *batches])
      main_sheet.append(["Image Size (HxW):", imgsz[0], imgsz[1]])
      main_sheet.append(["FP16 Export (half):", export_half])
      main_sheet.append(["FP16 Quantize:", quantize_fp16])
      main_sheet.append([])
      main_sheet.append(["System Information:"])
      try:
        main_sheet.append(["Hostname:", platform.node()])
        main_sheet.append(["OS:", platform.system()])
        main_sheet.append(["OS Version:", platform.version()])
        main_sheet.append(["OS Release:", platform.release()])
      except Exception as e:
        main_sheet.append([f"Cannot get OS information {e}"])
      main_sheet.append(["Python Version:", sys.version])
      main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
      main_sheet.append(["MIGraphX Version:", migraphx_version])
      main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)

      try:
        main_sheet.append(["CPU:", platform.processor()])
        import reports

        accelerators = reports.enumerate_accelerators()
        for item in accelerators.get("gpu", []):
          main_sheet.append(["GPU:", item.get("name")])
        for item in accelerators.get("npu", []):
          main_sheet.append(["NPU:", item.get("name")])
      except Exception as e:
        main_sheet.append([f"Cannot get accelerators information {e}"])

      try:
        main_sheet.append([])
        main_sheet.append(["Environment Variables:"])
        for key, value in sorted(os.environ.items(), key=lambda x: x[0]):
          main_sheet.append([key, value])
      except Exception as e:
        main_sheet.append([f"Cannot get environment variables {e}"])

      suffix = ""
      if export_half or quantize_fp16:
        suffix = "_fp16"
      workbook_name = f"{platform.node().lower()}_models.yolo11l.migx_driver_cache_gist{suffix}_{report_datetime.strftime('%Y%m%d_%H%M%S')}.xlsx"
      workbook_path = workbook_name

      reports_path = os.path.join(os.path.dirname(__file__), "..", "..", "reports", report_datetime.strftime("%Y%m%d"))
      if not os.path.exists(reports_path):
        os.makedirs(reports_path)
        try:
          from shutil import copy

          copy(
            os.path.join(os.path.dirname(__file__), "..", "..", "!StatViewer.xlsm"),
            os.path.join(reports_path, "!StatViewer.xlsm"),
          )
        except Exception as e:
          print(f'{{ "Error": "Failed to copy !StatViewer.xlsm {e}" }}')

      wb.save(workbook_path)
      os.replace(workbook_path, os.path.join(reports_path, workbook_path))
      print(f'{{ "Workbook": "{os.path.join(reports_path, workbook_path)}" }},')
    except Exception as e:
      print(f'{{ "Error": "Failed to generate report {e}" }},')
      import traceback

      traceback.print_exc()

  print('{ "Status": "Done" }')
  return 0


class Model(__import__("class_model", fromlist=["Model"]).Model):  # test_perf.py compatibility wrapper
  """
  `test_perf.py` expects a `Model` class in the imported module.

  This wrapper uses the gist-style export/compile path to create an `.mxr` cache,
  then runs inference via Python `migraphx` (like other MIGraphX models in this repo).
  """

  def __init__(self):
    super().__init__()
    self.model = None
    self.input_data = None

    self.imgsz = (640, 640)
    self.export_dynamic = True
    self.export_half = False
    self.quantize_fp16 = False
    self.force_driver_compile = False
    self.target = "gpu"  # can be changed by the harness if desired

    self.model_description = "YOLOv11l inference using gist-style MIGraphX export/compile + Python migraphx cache"

  def prepare_batch(self, batch_size: int):
    self.imgsz = (_round_up_to_stride(self.imgsz[0], 32), _round_up_to_stride(self.imgsz[1], 32))

    onnx_name = f"yolov11l_{batch_size}b_{self.imgsz[0]}x{self.imgsz[1]}_gist.onnx"
    onnx_path = self.get_file_path(onnx_name)
    mxr_path = onnx_path[:-4] + "mxr"

    ensure_onnx_exported(
      onnx_path=onnx_path,
      batch=batch_size,
      imgsz=self.imgsz,
      weights_path="yolov11l.pt",
      weights_url="https://github.com/ultralytics/assets/releases/download/v8.4.0/yolo11l.pt",
      export_half=self.export_half,
      export_dynamic=self.export_dynamic,
    )
    fix_onnx_input_shape_inplace(onnx_path, batch_size, self.imgsz)

    if not os.path.exists(mxr_path):
      compile_to_mxr(
        onnx_path=onnx_path,
        mxr_path=mxr_path,
        quantize_fp16=self.quantize_fp16,
        force_driver=self.force_driver_compile,
        target=self.target,
        allow_fallback=True,
      )

  def read(self):
    try:
      import migraphx  # type: ignore
    except Exception as e:
      raise Exception(
        "Python package `migraphx` is required for `test_perf.py` inference runs. "
        "If you only have `migraphx-driver`, run this module directly instead: "
        "`python -m models.yolo11l.migx_driver_cache_gist ...`"
      ) from e

    onnx_name = f"yolov11l_{self.batch_size}b_{self.imgsz[0]}x{self.imgsz[1]}_gist.onnx"
    onnx_path = self.get_file_path(onnx_name)
    mxr_path = onnx_path[:-4] + "mxr"
    self.model = migraphx.load(mxr_path)

  def prepare(self):
    import numpy as np

    self.input_data = np.random.randn(self.batch_size, 3, 640, 640).astype(np.float32)

  def inference(self):
    return self.model.run({"images": self.input_data})

  def shutdown(self):
    if self.model is not None:
      del self.model
      self.model = None


if __name__ == "__main__":
  raise SystemExit(main())

